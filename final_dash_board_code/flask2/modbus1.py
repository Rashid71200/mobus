import requests

from pymodbus.client.sync import ModbusSerialClient
from openpyxl import Workbook, load_workbook
from datetime import datetime
import time
import os

variable_value = 20
valueBackup = 0

def backup_modbus(address, l ):
    global valueBackup
    client = ModbusSerialClient(method="rtu", port="COM5", stopbits=1, bytesize=8, parity='N', baudrate=9600)
    try:
        client.connect()
        result = client.read_holding_registers(address=address, count=1, unit=(l+1))
        valueBackup = result.registers[0]
    except:
        time.sleep(1)
        backup_modbus(address, l )

    print("reading from backup....................................................................................................... ")
    return valueBackup


def start_modbus(variable_value):
    b = variable_value
    l_values = []  # List to store the values from each slave
    client = ModbusSerialClient(method="rtu", port="COM5", stopbits=1, bytesize=8, parity='N', baudrate=9600)
    
    try:
        client.connect()
    except:
        time.sleep(4)
        start_modbus(variable_value)

    for l in range(b):  # Iterate over the slave addresses (1 to b)
        print(f"number of loop {l}")
        valueS = []  # List to store the key-value pairs for each slave

        update_values = {
            "voltage": 3546,
            "current": 3654,
            "frequency": 3756,
            "power": 3878
        }
        
        filename = f"data{l+1}.xlsx"  # Use l in the filename for each slave

        if not os.path.exists(filename):
            workbook = Workbook()
            workbook.save(filename)
            
        workbook = load_workbook(filename)
        sheet = workbook.active
            
        for key, address in update_values.items():

            # Reading Holding Registers.............................
            try:
                result = client.read_holding_registers(address=address, count=1, unit=(l+1))
                value = result.registers[0]
            except:
                time.sleep(1)
                value = backup_modbus(address, l )
            #.......................................................


            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet.append([current_time, key, value])
            valueS.append({key: value})  # Format key and value as a dictionary

        workbook.save(filename)
            
        l_values.append(valueS)
        print('time to sleep')
        time.sleep(0.01)


    client.close()
    return l_values  # Return the list of dictionaries

def send_data_to_flask(value):
    url = 'http://localhost:5000/receive_data'  # Replace with the Flask server URL
    data = {'value': value}

    try:
        response = requests.post(url, json=data)
        if response.status_code == 200:
            print('Data sent successfully')
        else:
            print('Failed to send data')
    except requests.exceptions.RequestException as e:
        print(f'An error occurred: {e}')

while True:
    value = start_modbus(variable_value)
    send_data_to_flask(value)
    time.sleep(5)