from pymodbus.client.sync import ModbusSerialClient
from flask import Flask, render_template, request, session, redirect
from flask_socketio import join_room, leave_room, send, SocketIO
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import time
import os

variable_value = 20


def start_modbus(variable_value):
    b = variable_value
    l_values = []  # List to store the values from each slave
    client = ModbusSerialClient(method="rtu", port="COM5", stopbits=1, bytesize=8, parity='N', baudrate=9600)
    try:
        client.connect()
    except:
            time.sleep(1)

    for l in range(b):  # Iterate over the slave addresses (1 to b)
        print(f"number of loop {l}")
        valueS = []  # List to store the key-value pairs for each slave

        update_values = {
        "voltage": 3546,
        "current": 3654,
        "frequency": 3756,
        "power": 3878
        }
        filename = f"data{l+1}.xlsx"  # Use l in the filename for each slave
        try:
            #print("Connected")
            if not os.path.exists(filename):
                workbook = Workbook()
                workbook.save(filename)
            workbook = load_workbook(filename)
            sheet = workbook.active
            #print("Connectelk")
            for key, address in update_values.items():
                #print(f"Unit: {l+1}")
                result = client.read_holding_registers(address=address, count=1, unit=(l+1))
                value = result.registers[0]
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                sheet.append([current_time, key, value])
                valueS.append(f"{key}*{value}")  # Format key and value as a string
                #print(f"{key}: {value}"
            workbook.save(filename)
            l_values.append("$".join(valueS))
            print('time to sleep')
        except:
            time.sleep(1)


    client.close()
    return "$".join(l_values)  # Join all l_values to form the final string



app = Flask(__name__)
app.config["SECRET_KEY"] = "ahrneloy"
socketio = SocketIO(app)

@app.route("/", methods=["POST", "GET"])
def home():
    return render_template("styledashboard.html", variable_value=variable_value)

@socketio.on("update")
def update_voltage():

    value = start_modbus(variable_value)
    socketio.emit("update", {"value": value})

if __name__ == "__main__":
    socketio.run(app, debug=True)