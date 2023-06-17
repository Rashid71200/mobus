
from pymodbus.client.sync import ModbusSerialClient
from openpyxl import Workbook, load_workbook
from datetime import datetime
import time
import os

variable_value = 20


        

def start_modbus(variable_value):
    b = variable_value
    l_values = []  # List to store the values from each slave
    client = ModbusSerialClient(method="rtu", port="COM5", stopbits=1, bytesize=8, parity='N', baudrate=9600)
    
    try:
        client.connect()
    except:
        time.sleep(1)

    for l in range(b):  # Iterate over the slave addresses (1 to b)
        print(f"number of loop {l}")
        valueS = []  # List to store the key-value pairs for each slave

        update_values = {
            "voltage": 3546,
            "current": 3654,
            "frequency": 3756,
            "power": 3878
        }
        
        filename = f"data{l+1}.xlsx"  # Use l in the filename for each slave
        try:
            if not os.path.exists(filename):
                workbook = Workbook()
                workbook.save(filename)
            
            workbook = load_workbook(filename)
            sheet = workbook.active
            
            for key, address in update_values.items():
                result = client.read_holding_registers(address=address, count=1, unit=(l+1))
                value = result.registers[0]
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                sheet.append([current_time, key, value])
                valueS.append({key: value})  # Format key and value as a dictionary

            workbook.save(filename)
            
            l_values.append(valueS)
            print('time to sleep')
            time.sleep(0.2)
        except:
            time.sleep(1)

    client.close()
    return l_values  # Return the list of dictionaries

value = start_modbus(variable_value)
print(value)


[
    [{'voltage': 40}, {'current': 62}, {'frequency': 20}, {'power': 60}],
    [{'voltage': 55}, {'current': 25}, {'frequency': 51}, {'power': 42}],
    [{'voltage': 21}, {'current': 28}, {'frequency': 62}, {'power': 28}],
    [{'voltage': 31}, {'current': 44}, {'frequency': 55}, {'power': 76}],
    [{'voltage': 25}, {'current': 58}, {'frequency': 32}, {'power': 52}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0}, {'power': 0}], 
    [{'voltage': 0},{'current': 0}, {'frequency': 0}, {'power': 0}],
    [{'voltage': 0}, {'current': 0}, {'frequency': 0},{'power': 0}], 
    [{'voltage': 0}, {'current': 0},{'frequency': 0}, {'power': 0}]
]

